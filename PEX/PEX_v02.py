import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# 定义因素和水平
factors = {
    'Type': ['Two layers','Three layers' ],
    'Bot-Layer': ['NW STI', 'N+AA', 'N+Poly on AA', 'N+Poly on STI', 'M1', 'M2', 'M4', 'M5'],
    'Mid-Layer': ['M1', 'M2', 'M3', 'M4', 'M5', 'N+Poly'],
    'Top-Layer': ['M1', 'M2', 'M3', 'M4', 'M5', 'MT']
}

# 生成两层和两层的数据
two_layers_data = [
    ['Two layers', 'NW STI', 'M1', '/'],
    ['Two layers', 'NW STI', 'N+Poly', '/'],
    ['Two layers', 'N+AA', 'M1', '/'],
    ['Two layers', 'N+Poly on AA', 'M1', '/'],
    ['Two layers', 'N+Poly on STI', 'M1', '/'],
    ['Two layers', 'M1', 'M2', '/'],
    ['Two layers', 'M1', 'M3', '/'],
    ['Two layers', 'M2', 'M3', '/'],
    ['Two layers', 'M2', 'M4', '/'],
    ['Two layers', 'M5', 'MT', '/']
]

# 生成三层和两层的数据
three_layers_data = [
    ['Three layers', 'NW STI', 'M1', 'M2'],
    ['Three layers', 'NW STI', 'N+Poly', 'M1'],
    ['Three layers', 'N+AA', 'M1', 'M2'],
    ['Three layers', 'N+Poly on AA', 'M1', 'M2'],
    ['Three layers', 'N+Poly on STI', 'M1', 'M2'],
    ['Three layers', 'M1', 'M2', 'M3'],
    ['Three layers', 'M1', 'M2', 'M4'],
    ['Three layers', 'M1', 'M3', 'M4'],
    ['Three layers', 'M2', 'M3', 'M4'],
    ['Three layers', 'M2', 'M4', 'M5'],
    ['Three layers', 'M4', 'M5', 'MT']
]


# 合并数据
data = two_layers_data + three_layers_data

# 创建DataFrame
df = pd.DataFrame(data, columns=['Type', 'Bot-Layer', 'Mid-Layer', 'Top-Layer'])

# 定义最小线宽
Poly_W = '0.13'
M1_W = '0.16'
Mn_W = '0.2'
MT_W = '0.42'
Max_W = '10'

# 定义最小间距
Poly_S = '0.18'
M1_S = '0.17'
Mn_S = '0.2'
MT_S = '0.42'
Max1_S = '0.25'
Max2_S = '0.5'

# 定义最小线宽的倍数
Poly_W_value = float(Poly_W)  # 将字符串转换为浮点数
M1_W_value = float(M1_W)
Mn_W_value = float(Mn_W)
MT_W_value = float(MT_W)
Max_W_value = float(Max_W)

Poly_W_2x = str(round(Poly_W_value * 2, 2))  #计算2倍数的值并四舍五入到两位小数
Poly_W_3x = str(round(Poly_W_value * 3, 2))
M1_W_2x = str(round(M1_W_value * 2, 2))
M1_W_3x = str(round(M1_W_value * 3, 2))
Mn_W_2x= str(round(Mn_W_value * 2, 2))
Mn_W_3x = str(round(Mn_W_value * 3, 2))
MT_W_2x = str(round(MT_W_value * 2, 2))
MT_W_3x= str(round(MT_W_value * 3, 2))


# 定义最小间距的倍数
Poly_S_value = float(Poly_S)  # 将字符串转换为浮点数
M1_S_value = float(M1_S)
Mn_S_value = float(Mn_S)
MT_S_value = float(MT_S)
Max1_S_value = float(Max1_S)
Max2_S_value = float(Max2_S)  #

Poly_S_2x = str(round(Poly_S_value * 2, 2))  #计算2倍数的值并四舍五入到两位小数
Poly_S_3x = str(round(Poly_S_value * 3, 2))
M1_S_2x = str(round(M1_S_value * 2, 2))
M1_S_3x = str(round(M1_S_value * 3, 2))
Mn_S_2x= str(round(Mn_S_value * 2, 2))
Mn_S_3x= str(round(Mn_S_value * 3, 2))
MT_S_2x = str(round(MT_S_value * 2, 2))
MT_S_3x= str(round(MT_S_value * 3, 2))


# 定义 W 和 S 的水平
W_levels = [Poly_W, M1_W, Mn_W, MT_W, Max_W]
S_levels = [Poly_S, M1_S, Mn_S, MT_S, Max1_S, Max2_S]


# 遍历W_levels列表中的前四个数值
for start_value in W_levels[:4]:  # 只取前四个元素
    start_value = float(start_value)  # 将字符串转换为浮点数

    # 计算新值
    current_value_2x = round(start_value * 2, 2)  # 计算2倍数递增的值并四舍五入到两位小数
    current_value_3x = round(start_value * 3, 2)  # 计算3倍数递增的值并四舍五入到两位小数

    # 插入新值到当前元素后面插入新值
    index = W_levels.index(str(start_value)) + 1
    W_levels.insert(index, str(current_value_2x))  # 将新值2x插入到指定位置
    W_levels.insert(index + 1, str(current_value_3x))  # 将新值3x插入到指定位置

# 遍历S_levels列表中的前四个数值
for start_value in S_levels[:4]:  # 只取前四个元素
    start_value = float(start_value)  # 将字符串转换为浮点数

    # 计算新值
    current_value_2x = round(start_value * 2, 2)  # 计算2倍数递增的值并四舍五入到两位小数
    current_value_3x = round(start_value * 3, 2)  # 计算3倍数递增的值并四舍五入到两位小数

    # 插入新值到当前元素后面插入新值
    index = S_levels.index(str(start_value)) + 1
    S_levels.insert(index, str(current_value_2x))  # 将新值2x插入到指定位置
    S_levels.insert(index + 1, str(current_value_3x))  # 将新值3x插入到指定位置


# 生成所有可能的组合
expanded_data = []
for row in data:
    for w_level in W_levels:
        for s_level in S_levels:
            expanded_data.append(row + [w_level, s_level])

# 创建扩展后的DataFrame
expanded_df = pd.DataFrame(expanded_data, columns=['Type', 'Bot-Layer', 'Mid-Layer', 'Top-Layer', 'W', 'S'])

# 在第一列后增加两列Module和Label
expanded_df.insert(0, 'Module', '')
expanded_df.insert(1, 'Label', '')

# 在最后一列后面增加5列
expanded_df['NF'] = ''
expanded_df['L'] = ''
expanded_df['Row(Y)'] = ''
expanded_df['Col(X)'] = ''
expanded_df['Mid1'] = ''
expanded_df['Bottom'] = ''
expanded_df['Top'] = ''
expanded_df['Mid2'] = ''

# 定义文件名
file_name = 'PEX.xlsx'

# 定义有效的 W 和 S 组合
valid_three_layers_mt_combinations = [
    (MT_W, MT_S),
    (MT_W, MT_S_2x),
    (MT_W, MT_S_3x),
    (MT_W_2x, MT_S),
    (MT_W_3x, MT_S),
    (Max_W, Max2_S)
]

valid_three_layers_m5_combinations = [
    (Mn_W, Mn_S),
    (Mn_W, Mn_S_2x),
    (Mn_W, Mn_S_3x),
    (Mn_W_2x, Mn_S),
    (Mn_W_3x, Max1_S),
    (Max_W, Max2_S)
]

valid_three_layers_m4_m3_m2_combinations = [
    (Mn_W, Mn_S),
    (Mn_W, Mn_S_2x),
    (Mn_W, Mn_S_3x),
    (Mn_W_2x, Mn_S),
    (Mn_W_3x, Max1_S),
    (Max_W, Max2_S)
]

valid_three_layers_m1_combinations = [
    (M1_W, Poly_S),
    (M1_W, Poly_S_2x),
    (M1_W, Poly_S_3x),
    (M1_W_2x, Poly_S),
    (M1_W_3x, Mn_S),
    (Max_W, Max2_S)
]

valid_two_layers_mt_combinations = [
    (MT_W, MT_S),
    (MT_W, MT_S_2x),
    (MT_W, MT_S_3x),
    (MT_W_2x, MT_S),
    (MT_W_3x, MT_S),
    (Max_W, Max2_S)
]


valid_two_layers_m4_m3_m2_combinations = [
    (Mn_W, Mn_S),
    (Mn_W, Mn_S_2x),
    (Mn_W, Mn_S_3x),
    (Mn_W_2x, Mn_S),
    (Mn_W_3x, Max1_S),
    (Max_W, Max2_S)
]

valid_two_layers_m1_combinations = [
    (M1_W, M1_S),
    (M1_W, M1_S_2x),
    (M1_W, M1_S_3x),
    (M1_W_2x, M1_S),
    (M1_W_3x, Mn_S),
    (Max_W, Max2_S)
]

valid_two_layers_poly_combinations = [
    (Poly_W, Poly_S),
    (Poly_W, Poly_S_2x),
    (Poly_W, Poly_S_3x),
    (Poly_W_2x, Poly_S),
    (Poly_W_3x, Poly_S),
    (Max_W, Poly_S)
]

# 删除不满足条件的行
filtered_df = expanded_df[
    ~(
        ((expanded_df['Type'] == 'Three layers') &
         (expanded_df['Top-Layer'] == 'MT') &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_mt_combinations)))
        |
        ((expanded_df['Type'] == 'Three layers') &
         (expanded_df['Top-Layer'] == 'M5') &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_m5_combinations)))
        |
        ((expanded_df['Type'] == 'Three layers') &
         (expanded_df['Top-Layer'].isin(['M4', 'M3', 'M2'])) &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_m4_m3_m2_combinations)))
        |
        ((expanded_df['Type'] == 'Three layers') &
         (expanded_df['Top-Layer'] == 'M1') &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_m1_combinations)))
        |
        ((expanded_df['Type'] == 'Two layers') &
         (expanded_df['Mid-Layer'] == 'MT') &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_mt_combinations)))
        |
        ((expanded_df['Type'] == 'Two layers') &
         (expanded_df['Mid-Layer'].isin(['M4', 'M3', 'M2'])) &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_m4_m3_m2_combinations)))
        |
        ((expanded_df['Type'] == 'Two layers') &
         (expanded_df['Mid-Layer'] == 'M1') &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_m1_combinations)))
        |
        ((expanded_df['Type'] == 'Two layers') &
         (expanded_df['Mid-Layer'] == 'N+Poly') &
         (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_poly_combinations)))
    )
]

filtered_df.reset_index(drop=True, inplace=True)
filtered_df.insert(0, 'NO.', range(1, len(filtered_df) + 1))

# 保存筛选后的DataFrame到Excel文件
filtered_df.to_excel(file_name, index=False)

# 加载 Excel 文件
wb = load_workbook(file_name)

# 选择第一个工作表
ws = wb.active

# 设置字体和对齐方式
font = Font(name='Times New Roman')
alignment = Alignment(horizontal='center', vertical='center')

# 设置首行字体加粗
for cell in ws[1]:
    cell.font = Font(name='Times New Roman', bold=True)

# 遍历所有单元格，设置字体、对齐方式和行高
for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = font
        cell.alignment = alignment
    # 设置行高为 20
    ws.row_dimensions[row[0].row].height = 20

# 自动调整列宽以适应内容
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列名
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

# 保存格式化后的Excel文件
wb.save(file_name)

# 打开生成的Excel文件
os.startfile(file_name)

print(f"Excel文件已生成，并已保存为 '{file_name}'，并自动打开。")
