import sys
from PySide6.QtWidgets import QApplication, QWidget, QGridLayout, QLineEdit, QPushButton, QVBoxLayout, QLabel, QGraphicsBlurEffect
from PySide6.QtCore import Qt
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

class InputDialog(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('Input Dialog')
        self.setWindowOpacity(0.95)
        layout = QVBoxLayout()
        # 创建一个标题标签
        titleLabel = QLabel("请输入各层次的最小线宽和最小间距", self)
        titleLabel.setAlignment(Qt.AlignCenter)  # 设置居中对齐
        titleLabel.setStyleSheet("font-size: 18px; font-weight: bold; font-family: 'Microsoft YaHei', sans-serif;")  # 设置样式
        gridLayout = QGridLayout()
        self.labels = [
            QLabel("Poly_W:"), QLabel("M1_W:"), QLabel("Mn_W:"), QLabel("MT_W:"), QLabel("Max_W:"),
            QLabel("Poly_S:"), QLabel("M1_S:"), QLabel("Mn_S:"), QLabel("MT_S:"), QLabel("Max1_S:"), QLabel("Max2_S:")
        ]
        self.lineEdits = [
            QLineEdit(), QLineEdit(), QLineEdit(), QLineEdit(), QLineEdit(),
            QLineEdit(), QLineEdit(), QLineEdit(), QLineEdit(), QLineEdit(), QLineEdit()
        ]
        for i, (label, lineedit) in enumerate(zip(self.labels, self.lineEdits)):
            if i < 5:
                gridLayout.addWidget(label, i, 0)
                gridLayout.addWidget(lineedit, i, 1)
            else:
                gridLayout.addWidget(label, i - 5, 2)
                gridLayout.addWidget(lineedit, i - 5, 3)
        submitButton = QPushButton('Submit', self)
        submitButton.clicked.connect(self.showValues)
        self.resultLabel = QLabel(self)
        layout.addWidget(titleLabel)
        layout.addLayout(gridLayout)
        layout.addWidget(submitButton)
        layout.addWidget(self.resultLabel)
        self.setLayout(layout)
        self.setFixedSize(400, 600)
        self.setStyleSheet("""
            QWidget {
                background-color: #F0F0F0;
                border-radius: 15px;
                color: #000000;
                font-family: Arial, sans-serif;
            }
            QLineEdit {
                padding: 10px;
                border: 1px solid #A0A0A0;
                border-radius: 10px;
                background-color: rgba(255, 255, 255, 0.8);
                color: #000000;
                font-size: 16px;
            }
            QPushButton {
                background-color: #0078D7;
                color: white;
                border: none;
                border-radius: 10px;
                padding: 10px;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #005A9E;
            }
            QLabel {
                font-size: 16px;
                color: #000000;
                background-color: transparent;
            }
        """)
        blur_effect = QGraphicsBlurEffect()
        blur_effect.setBlurRadius(10)
        self.setGraphicsEffect(blur_effect)

    def showValues(self):
        global Poly_W, M1_W, Mn_W, MT_W, Max_W, Poly_S, M1_S, Mn_S, MT_S, Max1_S, Max2_S

        Poly_W = self.lineEdits[0].text()
        M1_W = self.lineEdits[1].text()
        Mn_W = self.lineEdits[2].text()
        MT_W = self.lineEdits[3].text()
        Max_W = self.lineEdits[4].text()
        Poly_S = self.lineEdits[5].text()
        M1_S = self.lineEdits[6].text()
        Mn_S = self.lineEdits[7].text()
        MT_S = self.lineEdits[8].text()
        Max1_S = self.lineEdits[9].text()
        Max2_S = self.lineEdits[10].text()

        # print(f"Poly_W: {Poly_W}, M1_W: {M1_W}, Mn_W: {Mn_W}, MT_W: {MT_W}, Max_W: {Max_W}")
        # print(f"Poly_S: {Poly_S}, M1_S: {M1_S}, Mn_S: {Mn_S}, MT_S: {MT_S}, Max1_S: {Max1_S}, Max2_S: {Max2_S}")

        self.generate_excel()

    def generate_excel(self):
        # 定义因素和水平，M=2
        factors = {
            'Type': ['Two layers', 'Three layers'],
            'Bot-Layer': ['NW STI', 'N+AA', 'N+Poly on AA', 'N+Poly on STI', 'M1'],
            'Mid-Layer': ['M1', 'N+Poly'],
            'Top-Layer': ['M1', 'MT']
        }

        two_layers_data = [
            ['Two layers', 'NW STI', 'M1', '/'],
            ['Two layers', 'NW STI', 'N+Poly', '/'],
            ['Two layers', 'N+AA', 'M1', '/'],
            ['Two layers', 'N+Poly on AA', 'M1', '/'],
            ['Two layers', 'N+Poly on STI', 'M1', '/']
        ]

        three_layers_data = [
            ['Three layers', 'NW STI', 'M1', 'MT'],
            ['Three layers', 'NW STI', 'N+Poly', 'M1'],
            ['Three layers', 'N+AA', 'M1', 'MT'],
            ['Three layers', 'N+Poly on AA', 'M1', 'MT'],
            ['Three layers', 'N+Poly on STI', 'M1', 'MT']
        ]
        # 定义因素和水平，M=3
        factors = {
            'Type': ['Two layers', 'Three layers'],
            'Bot-Layer': ['NW STI', 'N+AA', 'N+Poly on AA', 'N+Poly on STI', 'M1', 'M2'],
            'Mid-Layer': ['M1', 'M2', 'N+Poly'],
            'Top-Layer': ['M1', 'M2', 'MT']
        }

        two_layers_data = [
            ['Two layers', 'NW STI', 'M1', '/'],
            ['Two layers', 'NW STI', 'N+Poly', '/'],
            ['Two layers', 'N+AA', 'M1', '/'],
            ['Two layers', 'N+Poly on AA', 'M1', '/'],
            ['Two layers', 'N+Poly on STI', 'M1', '/'],
            ['Two layers', 'M1', 'M2', '/'],
            ['Two layers', 'M1', 'M3', '/'],
            ['Two layers', 'M2', 'M3', '/']
        ]

        three_layers_data = [
            ['Three layers', 'NW STI', 'M1', 'M2'],
            ['Three layers', 'NW STI', 'N+Poly', 'M1'],
            ['Three layers', 'N+AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on STI', 'M1', 'M2'],
            ['Three layers', 'M1', 'M2', 'M3']
        ]

        # 定义因素和水平，M=4
        factors = {
            'Type': ['Two layers', 'Three layers'],
            'Bot-Layer': ['NW STI', 'N+AA', 'N+Poly on AA', 'N+Poly on STI', 'M1', 'M2'],
            'Mid-Layer': ['M1', 'M2', 'M3', 'M4', 'N+Poly'],
            'Top-Layer': ['M1', 'M2', 'M3', 'MT']
        }

        two_layers_data = [
            ['Two layers', 'NW STI', 'M1', '/'],
            ['Two layers', 'NW STI', 'N+Poly', '/'],
            ['Two layers', 'N+AA', 'M1', '/'],
            ['Two layers', 'N+Poly on AA', 'M1', '/'],
            ['Two layers', 'N+Poly on STI', 'M1', '/'],
            ['Two layers', 'M1', 'M2', '/'],
            ['Two layers', 'M1', 'M3', '/'],
            ['Two layers', 'M2', 'M3', '/'],
            ['Two layers', 'M2', 'M4', '/']
           ]

        three_layers_data = [
            ['Three layers', 'NW STI', 'M1', 'M2'],
            ['Three layers', 'NW STI', 'N+Poly', 'M1'],
            ['Three layers', 'N+AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on STI', 'M1', 'M2'],
            ['Three layers', 'M1', 'M2', 'M3'],
            ['Three layers', 'M1', 'M2', 'M4'],
            ['Three layers', 'M1', 'M3', 'M4'],
            ['Three layers', 'M2', 'M3', 'M4'],
            ['Three layers', 'M2', 'M4', 'M5']
        ]

        # 定义因素和水平，M=5
        factors = {
            'Type': ['Two layers', 'Three layers'],
            'Bot-Layer': ['NW STI', 'N+AA', 'N+Poly on AA', 'N+Poly on STI', 'M1', 'M2'],
            'Mid-Layer': ['M1', 'M2', 'M3', 'M4', 'N+Poly'],
            'Top-Layer': ['M1', 'M2', 'M3', 'M4', 'MT']
        }

        two_layers_data = [
            ['Two layers', 'NW STI', 'M1', '/'],
            ['Two layers', 'NW STI', 'N+Poly', '/'],
            ['Two layers', 'N+AA', 'M1', '/'],
            ['Two layers', 'N+Poly on AA', 'M1', '/'],
            ['Two layers', 'N+Poly on STI', 'M1', '/'],
            ['Two layers', 'M1', 'M2', '/'],
            ['Two layers', 'M1', 'M3', '/'],
            ['Two layers', 'M2', 'M3', '/'],
            ['Two layers', 'M2', 'M4', '/']
        ]

        three_layers_data = [
            ['Three layers', 'NW STI', 'M1', 'M2'],
            ['Three layers', 'NW STI', 'N+Poly', 'M1'],
            ['Three layers', 'N+AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on STI', 'M1', 'M2'],
            ['Three layers', 'M1', 'M2', 'M3'],
            ['Three layers', 'M1', 'M2', 'M4'],
            ['Three layers', 'M1', 'M3', 'M4'],
            ['Three layers', 'M2', 'M3', 'M4'],
            ['Three layers', 'M2', 'M4', 'M5']
        ]
        # 定义因素和水平，M=6
        factors = {
            'Type': ['Two layers', 'Three layers'],
            'Bot-Layer': ['NW STI', 'N+AA', 'N+Poly on AA', 'N+Poly on STI', 'M1', 'M2', 'M4', 'M5'],
            'Mid-Layer': ['M1', 'M2', 'M3', 'M4', 'M5', 'N+Poly'],
            'Top-Layer': ['M1', 'M2', 'M3', 'M4', 'M5', 'MT']
        }

        two_layers_data = [
            ['Two layers', 'NW STI', 'M1', '/'],
            ['Two layers', 'NW STI', 'N+Poly', '/'],
            ['Two layers', 'N+AA', 'M1', '/'],
            ['Two layers', 'N+Poly on AA', 'M1', '/'],
            ['Two layers', 'N+Poly on STI', 'M1', '/'],
            ['Two layers', 'M1', 'M2', '/'],
            ['Two layers', 'M1', 'M3', '/'],
            ['Two layers', 'M2', 'M3', '/'],
            ['Two layers', 'M2', 'M4', '/'],
            ['Two layers', 'M5', 'MT', '/']
        ]

        three_layers_data = [
            ['Three layers', 'NW STI', 'M1', 'M2'],
            ['Three layers', 'NW STI', 'N+Poly', 'M1'],
            ['Three layers', 'N+AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on AA', 'M1', 'M2'],
            ['Three layers', 'N+Poly on STI', 'M1', 'M2'],
            ['Three layers', 'M1', 'M2', 'M3'],
            ['Three layers', 'M1', 'M2', 'M4'],
            ['Three layers', 'M1', 'M3', 'M4'],
            ['Three layers', 'M2', 'M3', 'M4'],
            ['Three layers', 'M2', 'M4', 'M5'],
            ['Three layers', 'M4', 'M5', 'MT']
        ]


        data = two_layers_data + three_layers_data
        df = pd.DataFrame(data, columns=['Type', 'Bot-Layer', 'Mid-Layer', 'Top-Layer'])

        # 定义最小线宽的倍数
        Poly_W_value = float(Poly_W)
        M1_W_value = float(M1_W)
        Mn_W_value = float(Mn_W)
        MT_W_value = float(MT_W)
        Max_W_value = float(Max_W)

        Poly_W_2x = str(round(Poly_W_value * 2, 2))
        Poly_W_3x = str(round(Poly_W_value * 3, 2))
        M1_W_2x = str(round(M1_W_value * 2, 2))
        M1_W_3x = str(round(M1_W_value * 3, 2))
        Mn_W_2x = str(round(Mn_W_value * 2, 2))
        Mn_W_3x = str(round(Mn_W_value * 3, 2))
        MT_W_2x = str(round(MT_W_value * 2, 2))
        MT_W_3x = str(round(MT_W_value * 3, 2))

        # 定义最小间距的倍数
        Poly_S_value = float(Poly_S)
        M1_S_value = float(M1_S)
        Mn_S_value = float(Mn_S)
        MT_S_value = float(MT_S)
        Max1_S_value = float(Max1_S)
        Max2_S_value = float(Max2_S)

        Poly_S_2x = str(round(Poly_S_value * 2, 2))
        Poly_S_3x = str(round(Poly_S_value * 3, 2))
        M1_S_2x = str(round(M1_S_value * 2, 2))
        M1_S_3x = str(round(M1_S_value * 3, 2))
        Mn_S_2x = str(round(Mn_S_value * 2, 2))
        Mn_S_3x = str(round(Mn_S_value * 3, 2))
        MT_S_2x = str(round(MT_S_value * 2, 2))
        MT_S_3x = str(round(MT_S_value * 3, 2))

        W_levels = [Poly_W, M1_W, Mn_W, MT_W, Max_W]
        S_levels = [Poly_S, M1_S, Mn_S, MT_S, Max1_S, Max2_S]

        for start_value in W_levels[:4]:
            start_value = float(start_value)
            current_value_2x = round(start_value * 2, 2)
            current_value_3x = round(start_value * 3, 2)
            index = W_levels.index(str(start_value)) + 1
            W_levels.insert(index, str(current_value_2x))
            W_levels.insert(index + 1, str(current_value_3x))

        for start_value in S_levels[:4]:
            start_value = float(start_value)
            current_value_2x = round(start_value * 2, 2)
            current_value_3x = round(start_value * 3, 2)
            index = S_levels.index(str(start_value)) + 1
            S_levels.insert(index, str(current_value_2x))
            S_levels.insert(index + 1, str(current_value_3x))

        expanded_data = []
        for row in data:
            for w_level in W_levels:
                for s_level in S_levels:
                    expanded_data.append(row + [w_level, s_level])

        expanded_df = pd.DataFrame(expanded_data, columns=['Type', 'Bot-Layer', 'Mid-Layer', 'Top-Layer', 'W', 'S'])
        expanded_df.insert(0, 'Module', '')
        expanded_df.insert(1, 'Label', '')
        expanded_df['NF'] = ''
        expanded_df['L'] = ''
        expanded_df['Row(Y)'] = ''
        expanded_df['Col(X)'] = ''
        expanded_df['Mid1'] = ''
        expanded_df['Bottom'] = ''
        expanded_df['Top'] = ''
        expanded_df['Mid2'] = ''

        file_name = 'expanded_layer_combinations_full_factorial.xlsx'

        valid_three_layers_mt_combinations = [
            (MT_W, MT_S), (MT_W, MT_S_2x), (MT_W, MT_S_3x), (MT_W_2x, MT_S), (MT_W_3x, MT_S), (Max_W, Max2_S)
        ]
        valid_three_layers_m5_combinations = [
            (Mn_W, Mn_S), (Mn_W, Mn_S_2x), (Mn_W, Mn_S_3x), (Mn_W_2x, Mn_S), (Mn_W_3x, Max1_S), (Max_W, Max2_S)
        ]
        valid_three_layers_m4_m3_m2_combinations = [
            (Mn_W, Mn_S), (Mn_W, Mn_S_2x), (Mn_W, Mn_S_3x), (Mn_W_2x, Mn_S), (Mn_W_3x, Max1_S), (Max_W, Max2_S)
        ]
        valid_three_layers_m1_combinations = [
            (M1_W, Poly_S), (M1_W, Poly_S_2x), (M1_W, Poly_S_3x), (M1_W_2x, Poly_S), (M1_W_3x, Mn_S), (Max_W, Max2_S)
        ]
        valid_two_layers_mt_combinations = [
            (MT_W, MT_S), (MT_W, MT_S_2x), (MT_W, MT_S_3x), (MT_W_2x, MT_S), (MT_W_3x, MT_S), (Max_W, Max2_S)
        ]
        valid_two_layers_m4_m3_m2_combinations = [
            (Mn_W, Mn_S), (Mn_W, Mn_S_2x), (Mn_W, Mn_S_3x), (Mn_W_2x, Mn_S), (Mn_W_3x, Max1_S), (Max_W, Max2_S)
        ]
        valid_two_layers_m1_combinations = [
            (M1_W, M1_S), (M1_W, M1_S_2x), (M1_W, M1_S_3x), (M1_W_2x, M1_S), (M1_W_3x, Mn_S), (Max_W, Max2_S)
        ]
        valid_two_layers_poly_combinations = [
            (Poly_W, Poly_S), (Poly_W, Poly_S_2x), (Poly_W, Poly_S_3x), (Poly_W_2x, Poly_S), (Poly_W_3x, Poly_S), (Max_W, Poly_S)
        ]

        filtered_df = expanded_df[
            ~(
                ((expanded_df['Type'] == 'Three layers') &
                 (expanded_df['Top-Layer'] == 'MT') &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_mt_combinations)))
                |
                ((expanded_df['Type'] == 'Three layers') &
                 (expanded_df['Top-Layer'] == 'M5') &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_m5_combinations)))
                |
                ((expanded_df['Type'] == 'Three layers') &
                 (expanded_df['Top-Layer'].isin(['M4', 'M3', 'M2'])) &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_m4_m3_m2_combinations)))
                |
                ((expanded_df['Type'] == 'Three layers') &
                 (expanded_df['Top-Layer'] == 'M1') &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_three_layers_m1_combinations)))
                |
                ((expanded_df['Type'] == 'Two layers') &
                 (expanded_df['Mid-Layer'] == 'MT') &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_mt_combinations)))
                |
                ((expanded_df['Type'] == 'Two layers') &
                 (expanded_df['Mid-Layer'].isin(['M4', 'M3', 'M2'])) &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_m4_m3_m2_combinations)))
                |
                ((expanded_df['Type'] == 'Two layers') &
                 (expanded_df['Mid-Layer'] == 'M1') &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_m1_combinations)))
                |
                ((expanded_df['Type'] == 'Two layers') &
                 (expanded_df['Mid-Layer'] == 'N+Poly') &
                 (~expanded_df[['W', 'S']].apply(tuple, axis=1).isin(valid_two_layers_poly_combinations)))
            )
        ]

        filtered_df.reset_index(drop=True, inplace=True)
        filtered_df.insert(0, 'NO.', range(1, len(filtered_df) + 1))

        filtered_df.to_excel(file_name, index=False)

        wb = load_workbook(file_name)
        ws = wb.active
        font = Font(name='Times New Roman')
        alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.font = Font(name='Times New Roman', bold=True)

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font
                cell.alignment = alignment
            ws.row_dimensions[row[0].row].height = 20

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_name)
        os.startfile(file_name)
        self.resultLabel.setText(f"Excel文件已生成，并已保存为 '{file_name}'，并自动打开。")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = InputDialog()
    ex.show()
    sys.exit(app.exec())
