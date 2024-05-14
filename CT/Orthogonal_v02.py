import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import tkinter as tk
from tkinter import messagebox, ttk
import os


def generate_experiment_design(variables):
    # 添加标题项到组合列表中
    all_combinations = [list(variables.keys())]
    all_combinations.extend(list(product(*variables.values())))

    # 将组合转换为DataFrame
    design_df = pd.DataFrame(all_combinations[1:], columns=all_combinations[0])

    # 在 DataFrame 中插入一列序号
    design_df.insert(0, 'NO.', range(1, len(design_df) + 1))

    # 在第一列后增加两列Module和Lable
    design_df.insert(1, 'Module', '')
    design_df.insert(2, 'Lable', '')

    # 在最后一列后面增加5列Sens1，Force，COM，Sens2，Description
    design_df['Sens1'] = ''
    design_df['Force'] = ''
    design_df['COM'] = ''
    design_df['Sens2'] = ''
    design_df['Description'] = ''

    # 保存设计到Excel文件
    excel_file = "orthogonal_experiment_design_allpairspy.xlsx"

    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 添加标题行并设置边框
    for c_idx, title in enumerate(design_df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=title)
        cell.font = Font(name='Times New Roman', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='medium'))

    # 将DataFrame写入工作表
    for r_idx, r in enumerate(design_df.values.tolist(), 2):
        for c_idx, value in enumerate(r, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman')

    # 设置字体和自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 设置行高
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            ws.row_dimensions[cell.row].height = 25

    # 保存Excel文件
    wb.save(excel_file)
    messagebox.showinfo("提示", f"正交实验设计已保存到 {excel_file}")

    # 自动打开Excel文件
    os.startfile(excel_file)


def generate_design():
    variables = {}
    for entry in variable_entries:
        variable_name = entry[0].get()
        variable_input = entry[1].get()
        if not variable_name or not variable_input:
            messagebox.showerror("错误", "变量名和水平不能为空")
            return
        variables[variable_name] = variable_input.split(',')
    generate_experiment_design(variables)


def add_variable_entry():
    if len(variable_entries) >= 8:
        messagebox.showwarning("警告", "已达到最大变量数")
        return

    variable_frame = tk.Frame(root, bg="white")
    variable_frame.pack(fill=tk.X, padx=10, pady=5)

    variable_label = tk.Label(variable_frame, text="变量名：", width=10, anchor='w', bg="white",
                              font=("Times New Roman", 10))
    variable_label.pack(side=tk.LEFT)

    variable_entry = tk.Entry(variable_frame, width=20, bg="white", font=("Times New Roman", 10))
    variable_entry.pack(side=tk.LEFT, padx=5)

    level_label = tk.Label(variable_frame, text="水平：", width=10, anchor='w', bg="white", font=("Times New Roman", 10))
    level_label.pack(side=tk.LEFT)

    level_entry = tk.Entry(variable_frame, width=50, bg="white", font=("Times New Roman", 10))
    level_entry.pack(side=tk.LEFT, padx=5)

    variable_entries.append((variable_entry, level_entry))


root = tk.Tk()
root.title("生成正交实验设计")
root.geometry("600x400")
root.config(bg="white")

title_label = tk.Label(root, text="请输入实验变量及其水平", font=("Arial", 12), bg="white")
title_label.pack(pady=10)

variable_entries = []

variable_frame = tk.Frame(root, bg="white")
variable_frame.pack(fill=tk.X, padx=10, pady=5)
add_variable_entry()

add_variable_button = tk.Button(root, text="添加变量", command=add_variable_entry, bg="#d9d9d9",
                                font=("Times New Roman", 10))
add_variable_button.pack(pady=5)

generate_button = tk.Button(root, text="生成设计", command=generate_design, bg="#cfe7d1", font=("Times New Roman", 12))
generate_button.pack(pady=10)

root.mainloop()
