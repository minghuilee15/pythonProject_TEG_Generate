import tkinter as tk
from tkinter import ttk
import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from ttkbootstrap import Style
import os

def generate_design():
    # 从界面获取用户输入的变量及其水平


    variables = {}
    for label, combobox in zip(variables_labels, variables_comboboxes):
        levels = combobox.get().split(',')
        variables[label] = [level.strip() for level in levels]

    # 添加标题项到组合列表中
    all_combinations = [list(variables.keys())]
    all_combinations.extend(list(product(*variables.values())))

    # 将组合转换为DataFrame
    design_df = pd.DataFrame(all_combinations[1:], columns=all_combinations[0])

    # 在第一列后增加两列Module和Label
    design_df.insert(1, 'Module', '')
    design_df.insert(2, 'Label', '')

    # 在最后一列后面增加5列Sens1，Force，COM，Sens2，Description
    design_df['Sens1'] = ''
    design_df['Force'] = ''
    design_df['COM'] = ''
    design_df['Sens2'] = ''
    design_df['Description'] = ''

    #定义Typical_CT尺寸
    Typical_CT_Size = '0.16'

    #拉偏CT_Size时候，控制Enclosure1_Size大小不变
    Enclosure1_Size = '0.06'

    # 拉偏CT_Size时候，控制Enclosure2_Size大小不变
    Enclosure2_Size_4T_Chain = '0'
    Enclosure2_Size_4T_Chain_End = '0.05'

    # 拉偏AA enc CT时候，控制Enclosure1_Size为0和0.03，Enclosure2_Size为0(4T,Chain)
    AA_E1_1 = '0'
    AA_E1_2 = '0.03'
    AA_E2_1 = '0' #4T,Chain
    AA_E2_2 = '0.05'  #4T_End,Chain_End

    # 拉偏M1 enc CT时候，控制Enclosure1_Size为0.06，Enclosure2_Size为-0.01(4T,Chain)，0和0.03(4T_End,Chain_End)
    M1_E1_1 = '0.06'
    M1_E2_1 = '-0.01' #4T,Chain
    M1_E2_2 = '0'  #4T_End,Chain_End
    M1_E2_3 = '0.03'  #4T_End,Chain_End

    # 拉偏Poly enc CT时候，控制Enclosure1_Size为0和0.03，Enclosure2_Size为0(4T,Chain)和0.05(4T_End,Chain_End)
    Poly_E1_1 = '0'
    Poly_E1_2 = '0.03'
    Poly_E2_1 = '0'#4T,Chain
    Poly_E2_2 = '0.05'  #4T_End,Chain_End

    # 删除不满足条件的行
    design_df = design_df[
        ~(
                # 删除AA因素中的Poly enc CT
                ((design_df['Device3'] == 'AA') & (design_df['Split'] == 'Poly enc CT')) |

                # 删除Poly因素中的AA enc CT
                ((design_df['Device3'] == 'Poly') & (design_df['Split'] == 'AA enc CT')) |


                # 定义Typical_CT_Size
                ((design_df['Split'] == 'AA enc CT') & (design_df['Size'] != Typical_CT_Size)) |
                ((design_df['Split'] == 'M1 enc CT') & (design_df['Size'] != Typical_CT_Size)) |
                ((design_df['Split'] == 'Poly enc CT') & (design_df['Size'] != Typical_CT_Size)) |

                # 拉偏CT_Size时候，控制Enclosure1_Size大小不变
                ((design_df['Split'] == 'Size') & (design_df['Enclosure1'] != Enclosure1_Size)) |

                # 拉偏CT_Size时候，控制Enclosure2_Size大小不变
                ((design_df['Type'] == '4T') & (design_df['Split'] == 'Size') & ~(design_df['Enclosure2'] == Enclosure2_Size_4T_Chain)) |
                ((design_df['Type'] == '4T-End') & (design_df['Split'] == 'Size') & ~(design_df['Enclosure2'] == Enclosure2_Size_4T_Chain_End)) |
                ((design_df['Type'] == 'Chain') & (design_df['Split'] == 'Size') & ~(design_df['Enclosure2'] == Enclosure2_Size_4T_Chain)) |
                ((design_df['Type'] == 'Chain-End') & (design_df['Split'] == 'Size') & ~(design_df['Enclosure2'] == Enclosure2_Size_4T_Chain_End)) |


                # AA enc CT

                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == '4T') & (~(design_df['Enclosure1'].isin([AA_E1_1, AA_E1_2])))) |
                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == '4T') & (design_df['Enclosure2'] != AA_E2_1)) |

                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == '4T-End') & (~(design_df['Enclosure1'].isin([AA_E1_1, AA_E1_2])))) |
                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == '4T-End') & (design_df['Enclosure2'] != AA_E2_2)) |

                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == 'Chain') & (~(design_df['Enclosure1'].isin([AA_E1_1, AA_E1_2])))) |
                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == 'Chain') & (design_df['Enclosure2'] != AA_E2_1)) |

                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == 'Chain-End') & (~(design_df['Enclosure1'].isin([AA_E1_1, AA_E1_2])))) |
                ((design_df['Split'] == 'AA enc CT') & (design_df['Type'] == 'Chain-End') & (design_df['Enclosure2'] != AA_E2_2)) |

                # M1 enc CT

                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == '4T') & (design_df['Enclosure1'] != M1_E1_1)) |
                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == '4T') & (design_df['Enclosure2'] != M1_E2_1)) |

                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == '4T-End') & ~(design_df['Enclosure1'] == M1_E1_1)) |
                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == '4T-End') & (~(design_df['Enclosure2'].isin([M1_E2_2, M1_E2_3])))) |

                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == 'Chain') & ~(design_df['Enclosure1'] == M1_E1_1)) |
                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == 'Chain') & (design_df['Enclosure2'] != M1_E2_1)) |

                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == 'Chain-End') & ~(design_df['Enclosure1'] == M1_E1_1)) |
                ((design_df['Split'] == 'M1 enc CT') & (design_df['Type'] == 'Chain-End') & (~(design_df['Enclosure2'].isin([M1_E2_2, M1_E2_3])))) |

                # Poly enc CT

                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == '4T') & (~(design_df['Enclosure1'].isin([Poly_E1_1, Poly_E1_2])))) |
                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == '4T') & (design_df['Enclosure2'] != Poly_E2_1)) |

                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == '4T-End') & (~(design_df['Enclosure1'].isin([Poly_E1_1, Poly_E1_2])))) |
                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == '4T-End') & (design_df['Enclosure2'] != Poly_E2_2)) |

                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == 'Chain') & (~(design_df['Enclosure1'].isin([Poly_E1_1, Poly_E1_2])))) |
                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == 'Chain') & (design_df['Enclosure2'] != Poly_E2_1)) |

                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == 'Chain-End') & (~(design_df['Enclosure1'].isin([Poly_E1_1, Poly_E1_2])))) |
                ((design_df['Split'] == 'Poly enc CT') & (design_df['Type'] == 'Chain-End') & (design_df['Enclosure2'] != Poly_E2_2))


        )
    ]

    # 在 DataFrame 中插入一列序号
    design_df.insert(0, 'NO.', range(1, len(design_df) + 1))

    # 保存设计到Excel文件
    excel_file = "orthogonal_experiment_design_CT.xlsx"

    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active

    # 添加标题行并设置边框
    for c_idx, title in enumerate(design_df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=title)
        cell.font = Font(name='Times New Roman', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(bottom=Side(style='medium'))

    # 将DataFrame写入工作表
    for r_idx, r in enumerate(design_df.values.tolist(), 2):
        for c_idx, value in enumerate(r, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman')

    # 设置字体和自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # 设置行高
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            ws.row_dimensions[cell.row].height = 25

    # 保存Excel文件
    wb.save(excel_file)
    print(f"正交实验设计已保存到 {excel_file}")

    # 自动打开Excel文件
    os.system(f'start excel {excel_file}')

# 创建主窗口
root = tk.Tk()
root.title("orthogonal_experiment_design")

# 设置主题
style = Style(theme='flatly')

# 设置字体
style.configure('TLabel', font=('Times New Roman', 12, 'bold'))
style.configure('TButton', font=('Microsoft YaHei', 12))

# 添加标签和输入框
variables_frame = ttk.LabelFrame(root, text="实验变量及水平")
variables_frame.grid(row=0, column=0, padx=200, pady=50, sticky="nsew")
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

variables_labels = ['Device', 'Device2', 'Device3', 'Type', 'Split', 'Size', 'Enclosure1', 'Enclosure2']
variables_comboboxes = []

# 设置每个部件的行和列权重
for i in range(len(variables_labels)):
    variables_frame.grid_columnconfigure(i * 2, weight=100)
    variables_frame.grid_columnconfigure(i * 2 + 1, weight=100)

for idx, label in enumerate(variables_labels):
    ttk.Label(variables_frame, text=label + ":").grid(row=idx, column=0, padx=5, pady=2)
    # 使用ttkbootstrap内置的输入框样式，并设置线宽
    combobox = ttk.Entry(variables_frame, style='secondary.TEntry', width=40)
    combobox.grid(row=idx, column=1, padx=5, pady=2)
    style.configure('secondary.TEntry', borderwidth=20)  # 设置输入框的线宽为2像素
    variables_comboboxes.append(combobox)

# 添加生成按钮
generate_button = ttk.Button(root, text="生成TEG表单", command=generate_design)
generate_button.grid(row=1, column=0, pady=10)

root.mainloop()



