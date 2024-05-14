import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os

# 定义实验变量及其水平
variables = {
    'Device': ['CT'],
    'Device2': ['N+', 'P+'],
    'Device3': ['AA', 'Poly'],
    'Type': ['4T', '4T-End', 'Chain', 'Chain-End'],
    'Split': ['Size', 'AA enc CT','M1 enc CT','Poly enc CT'],
    'Size': ['0.16', '0.144'],
    'Enclosure1': ['0.06', '0.03', '0'],
    'Enclosure2': ['0', '-0.01'],
}

# 添加标题项到组合列表中
all_combinations = [list(variables.keys())]
all_combinations.extend(list(product(*variables.values())))

# 将组合转换为DataFrame
design_df = pd.DataFrame(all_combinations[1:], columns=all_combinations[0])

# 在 DataFrame 中插入一列序号
design_df.insert(0, 'NO.', range(1, len(design_df) + 1))

# 在第一列后增加两列Module和Lable
design_df.insert(1, 'Module', '')
design_df.insert(2, 'Lable', '')

# 在最后一列后面增加5列Sens1，Force，COM，Sens2，Description
design_df['Sens1'] = ''
design_df['Force'] = ''
design_df['COM'] = ''
design_df['Sens2'] = ''
design_df['Description'] = ''

# 保存设计到Excel文件
excel_file = "orthogonal_experiment_design_allpairspy.xlsx"

# 创建一个新的Excel工作簿
wb = Workbook()
ws = wb.active

# 添加标题行并设置边框
for c_idx, title in enumerate(design_df.columns, 1):
    cell = ws.cell(row=1, column=c_idx, value=title)
    cell.font = Font(name='Times New Roman', bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(bottom=Side(style='medium'))

# 将DataFrame写入工作表
for r_idx, r in enumerate(design_df.values.tolist(), 2):
    for c_idx, value in enumerate(r, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(name='Times New Roman')

# 设置字体和自动调整列宽
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width

# 设置行高
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for cell in row:
        ws.row_dimensions[cell.row].height = 25

# 保存Excel文件
wb.save(excel_file)
print(f"正交实验设计已保存到 {excel_file}")

# 自动打开Excel文件
os.system(f'start excel {excel_file}')
